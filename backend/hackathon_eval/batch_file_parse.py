"""Parse bulk repo lists from CSV or Excel (first sheet)."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from pydantic import BaseModel, Field

_REPO_HEADERS = frozenset({"repo_url", "url", "repository", "repo", "github", "github_url"})
_BRANCH_HEADERS = frozenset({"branch", "ref", "git_ref"})


def _norm_header(h: str) -> str:
    s = (h or "").strip().lower().replace(" ", "_")
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    return s.strip("_") or ""


class BatchEvaluateItem(BaseModel):
    repo_url: str | None = None
    branch: str | None = None
    submission_context: str | None = None
    document_text: str | None = None
    submission_metadata: dict[str, Any] | None = None
    label: str | None = Field(default=None)

    model_config = {"extra": "ignore"}


def _row_to_item(row: dict[str, str], row_index: int) -> BatchEvaluateItem | None:
    norm = {_norm_header(k): (v or "").strip() for k, v in row.items() if k}
    url = ""
    for key in _REPO_HEADERS:
        if norm.get(key):
            url = norm[key]
            break
    if not url:
        return None
    branch = ""
    for key in _BRANCH_HEADERS:
        if norm.get(key):
            branch = norm[key]
            break
    reserved = _REPO_HEADERS | _BRANCH_HEADERS
    meta: dict[str, Any] = {}
    for k, v in norm.items():
        if k in reserved or v == "":
            continue
        meta[k] = v
    return BatchEvaluateItem(
        repo_url=url,
        branch=branch or None,
        submission_metadata=meta if meta else None,
        label=str(row_index),
    )


def parse_csv_bytes(data: bytes) -> list[BatchEvaluateItem]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    items: list[BatchEvaluateItem] = []
    for i, raw in enumerate(reader):
        row = {str(k or ""): str(v or "") for k, v in raw.items()}
        item = _row_to_item(row, i + 1)
        if item:
            items.append(item)
    return items


def parse_xlsx_bytes(data: bytes) -> list[BatchEvaluateItem]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for .xlsx; pip install openpyxl") from e

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        headers = [_norm_header(str(h) if h is not None else "") for h in header_row]
        data_rows: list[dict[str, str]] = []
        for row in rows:
            d: dict[str, str] = {}
            for j, cell in enumerate(row):
                if j >= len(headers) or not headers[j]:
                    continue
                val = "" if cell is None else str(cell).strip()
                d[headers[j]] = val
            if any(v.strip() for v in d.values()):
                data_rows.append(d)
    finally:
        wb.close()

    items: list[BatchEvaluateItem] = []
    for i, row in enumerate(data_rows):
        item = _row_to_item(row, i + 1)
        if item:
            items.append(item)
    return items


def parse_batch_upload(filename: str, data: bytes) -> list[BatchEvaluateItem]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv_bytes(data)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(data)
    raise ValueError("Unsupported file type: use .csv or .xlsx")
